"""
Structured incident feeds (V2 incident layer, tier 2): government-reported
spills merged into `incident_reports` alongside the news monitor.

Sources:
  NRC — National Response Center (ACTIVE). Federally mandated reporting of
  significant releases, national coverage, weekly-updated FOIA workbook
  (https://nrc.uscg.mil/FOIAFiles/CY{yy}.xlsx). Sewage/wastewater releases
  are filtered and matched against Medium/Large qualified targets by
  city+state. Because reporting is a legal requirement, this catches
  incidents that never make the news.

  State portals (STUBS — phase-next adapters):
    TX TCEQ spills, CA CIWQS SSO, FL DEP Public Notice of Pollution.
    All three are interactive web apps without stable machine endpoints
    (verified 2026-07-03: CIWQS = session form servlet, FL PNP = 302 to
    session app, TCEQ = no Socrata dataset). Each needs its own
    reverse-engineering pass; the adapter interface below is ready for
    them. NRC covers all three states in the meantime.

Usage:
    python structured_incidents.py            # NRC pull + merge
    python structured_incidents.py --dry-run
"""

import argparse
import hashlib
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone

import openpyxl
import requests
from dotenv import load_dotenv
from google.cloud import bigquery

load_dotenv()
load_dotenv(os.path.expanduser("~/.config/ipi-etl/.env"))  # secrets live outside the synced repo dir

CACHE_DIR = os.path.join(os.path.dirname(__file__), "data_cache")
NRC_URL_TEMPLATE = "https://nrc.uscg.mil/FOIAFiles/CY{yy}.xlsx"
LOOKBACK_DAYS = 30

SEWAGE_RE = re.compile(r"sewage|sewer|wastewater|effluent|sanitary", re.I)


def _norm_city(c: str) -> str:
    return re.sub(r"[^a-z ]", "", (c or "").lower()).strip()


def _parse_nrc_datetime(raw):
    """NRC workbook dates arrive as 'M/D/YYYY H:MM' strings (sometimes as
    real datetime cells). Returns tz-aware UTC datetime or None."""
    if isinstance(raw, datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=timezone.utc)
    s = str(raw or "").strip()
    for fmt in ("%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def fetch_targets(client, project_id):
    sql = f"""
    SELECT municipality_key, city, state
    FROM `{project_id}.ipi_intelligence.qualified_targets`
    WHERE size_tier IN ('Medium', 'Large') AND city IS NOT NULL
    """
    return {
        (_norm_city(r.city), r.state): (r.municipality_key, r.city)
        for r in client.query(sql).result()
    }


# ---------------------------------------------------------------------------
# NRC adapter
# ---------------------------------------------------------------------------

def fetch_nrc(targets: dict) -> list[dict]:
    """Download the current-year NRC workbook and return sewage incidents
    at qualified municipalities within the lookback window."""
    now = datetime.now(timezone.utc)
    os.makedirs(CACHE_DIR, exist_ok=True)
    path = os.path.join(CACHE_DIR, "nrc_current.xlsx")

    # Re-download if cache older than 1 day
    fresh = (os.path.exists(path)
             and time.time() - os.path.getmtime(path) < 86400)
    if not fresh:
        url = NRC_URL_TEMPLATE.format(yy=now.strftime("%y"))
        print(f"Downloading NRC workbook: {url}")
        resp = requests.get(url, timeout=300)
        if resp.status_code != 200 and now.month == 1:
            # January edge: current-year file may not exist yet
            url = NRC_URL_TEMPLATE.format(yy=(now.year - 1) % 100)
            resp = requests.get(url, timeout=300)
        resp.raise_for_status()
        with open(path, "wb") as f:
            f.write(resp.content)
    else:
        print("Using cached NRC workbook (<1 day old)")

    wb = openpyxl.load_workbook(path, read_only=True)

    # Materials per SEQNOS (sheet name has a known typo variant)
    materials = {}
    mat_sheet = next((s for s in ("MATERIAL_INVOLVED", "MATERIAL_INV0LVED")
                      if s in wb.sheetnames), None)
    if mat_sheet:
        rows = wb[mat_sheet].iter_rows(values_only=True)
        headers = [str(h or "").upper() for h in next(rows)]
        try:
            i_seq = headers.index("SEQNOS")
            i_name = next(i for i, h in enumerate(headers) if "MATERIAL" in h or "NAME" in h)
            for r in rows:
                seq = str(r[i_seq] or "")
                if seq:
                    materials.setdefault(seq, []).append(str(r[i_name] or ""))
        except (ValueError, StopIteration):
            pass

    ws = wb["INCIDENT_COMMONS"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "").upper() for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    cutoff = now - timedelta(days=LOOKBACK_DAYS)
    out, scanned, sewage_n = [], 0, 0
    for r in rows:
        scanned += 1
        desc = str(r[idx["DESCRIPTION_OF_INCIDENT"]] or "")
        seq = str(r[idx["SEQNOS"]] or "")
        mats = " ".join(materials.get(seq, []))
        if not (SEWAGE_RE.search(desc) or SEWAGE_RE.search(mats)):
            continue
        sewage_n += 1

        city = str(r[idx["LOCATION_NEAREST_CITY"]] or "")
        state = str(r[idx["LOCATION_STATE"]] or "").strip().upper()
        key = (_norm_city(city), state)
        if key not in targets:
            continue

        raw_dt = r[idx["INCIDENT_DATE_TIME"]]
        incident_dt = _parse_nrc_datetime(raw_dt)
        if incident_dt is None or incident_dt < cutoff:
            continue

        municipality_key, target_city = targets[key]
        headline = f"NRC #{seq}: {desc[:300]}" if desc else f"NRC report #{seq}"
        out.append({
            "incident_id": hashlib.sha1(f"NRC-{seq}".encode()).hexdigest(),
            "municipality_key": municipality_key,
            "city": target_city,
            "state": state,
            "headline": headline[:500],
            "url": "https://nrc.uscg.mil/",
            "news_source": "National Response Center",
            "incident_type": "sewer_overflow",
            "published_at": incident_dt.isoformat(),
            "first_seen_at": now.isoformat(),
        })

    print(f"NRC: {scanned:,} incidents scanned | {sewage_n} sewage-related | "
          f"{len(out)} at qualified targets within {LOOKBACK_DAYS}d")
    return out


# Adapter registry — add state adapters here as their endpoints get built.
ADAPTERS = {
    "nrc": fetch_nrc,
    # "tx_tceq": fetch_tx_tceq,     # interactive portal — needs RE pass
    # "ca_ciwqs": fetch_ca_ciwqs,   # session servlet — needs RE pass
    # "fl_pnp": fetch_fl_pnp,       # session app — needs RE pass
}


def merge_incidents(client, project_id, rows: list[dict]):
    schema = [
        bigquery.SchemaField("incident_id", "STRING"),
        bigquery.SchemaField("municipality_key", "STRING"),
        bigquery.SchemaField("city", "STRING"),
        bigquery.SchemaField("state", "STRING"),
        bigquery.SchemaField("headline", "STRING"),
        bigquery.SchemaField("url", "STRING"),
        bigquery.SchemaField("news_source", "STRING"),
        bigquery.SchemaField("incident_type", "STRING"),
        bigquery.SchemaField("published_at", "TIMESTAMP"),
        bigquery.SchemaField("first_seen_at", "TIMESTAMP"),
    ]
    temp_id = f"{project_id}.ipi_intelligence._temp_struct_{int(time.time())}"
    client.create_table(bigquery.Table(temp_id, schema=schema))
    errors = client.insert_rows_json(temp_id, rows)
    if errors:
        print(f"Insert errors: {errors[:3]}", file=sys.stderr)
        client.delete_table(temp_id, not_found_ok=True)
        sys.exit(1)
    job = client.query(f"""
    MERGE `{project_id}.ipi_intelligence.incident_reports` T
    USING `{temp_id}` S
    ON T.incident_id = S.incident_id
    WHEN NOT MATCHED THEN
      INSERT (incident_id, municipality_key, city, state, headline, url,
              news_source, incident_type, published_at, first_seen_at, dismissed)
      VALUES (S.incident_id, S.municipality_key, S.city, S.state, S.headline,
              S.url, S.news_source, S.incident_type, S.published_at,
              S.first_seen_at, FALSE)
    """)
    job.result()
    client.delete_table(temp_id, not_found_ok=True)
    print(f"New structured incidents merged: {job.num_dml_affected_rows or 0}")


def main():
    parser = argparse.ArgumentParser(description="Structured incident feeds")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    project_id = os.getenv("GCP_PROJECT_ID", "ipi-consent-decree-dashboard")
    client = bigquery.Client(project=project_id)

    from incident_monitor import INCIDENTS_DDL
    client.query(INCIDENTS_DDL.format(project=project_id)).result()

    targets = fetch_targets(client, project_id)
    print(f"Qualified Medium/Large targets: {len(targets)}")

    all_rows = []
    for name, adapter in ADAPTERS.items():
        try:
            all_rows.extend(adapter(targets))
        except Exception as exc:
            print(f"Adapter {name} failed: {exc}", file=sys.stderr)

    if args.dry_run or not all_rows:
        for r in all_rows[:10]:
            print(f"  [{'dry-run' if args.dry_run else 'found'}] "
                  f"{r['city']}, {r['state']} | {r['headline'][:90]}")
        if args.dry_run:
            return
    if all_rows:
        merge_incidents(client, project_id, all_rows)


if __name__ == "__main__":
    main()
